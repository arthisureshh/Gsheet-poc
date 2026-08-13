"""
Generate stress_test.xlsx — one file covering all hard detection scenarios.

Sheet1 "Mixed Layout"
  - Vertical table: Employee Performance (cols A-F, 50 data rows)
    Title row: plain text, light-blue bg, NO bold
    Header row: plain text, grey bg, NO bold
  - Horizontal table: Monthly KPIs (cols H-R, metrics as rows, months as cols)
    Title row: plain text, orange bg, NO bold
    Header row: plain text, yellow bg, NO bold
  - Single empty column (G) between the two tables
  - 8 accidental stray text cells scattered outside both tables

Sheet2 "Dense Tables"
  - Table A: Project Tasks (rows 1-31, cols A-G)
  - Table B: Budget Tracker (rows 1-31, cols I-N)  ← same row range, 1-col gap at H
  - Table C: Risk Register (rows 33-63, cols A-N)  ← 1 empty row gap
  All titles/headers: plain text, color-only (no bold, no merge)

Sheet3 "Random Scatter"
  - 4 tables at non-adjacent positions
  - Table 1: Customer Orders (A3:H43) — plain header, green bg
  - Table 2: Support Tickets (K3:P33) — plain header, purple bg
  - Table 3: KPI Matrix horizontal (A47:I57) — plain header, teal bg
  - Table 4: Product Lookup (K47:N67) — plain header, pink bg
  - Stray text in 5 cells between tables

Sheet4 "Attendance"
  - 25 team members x 60 work days
  - Row headers = member names (plain, no bold)
  - Column headers = dates (plain, no bold, light grey bg)
  - Data cells: GREEN fill = present, RED fill = absent, empty string value

Run:
    python create_stress_test.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

random.seed(99)

# ── Color palette (all plain — no bold anywhere in this file) ─────────────────
def fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

C_TITLE_BLUE   = fill("BDD7EE")   # light blue  — Sheet1 vertical title
C_HDR_GREY     = fill("D9D9D9")   # grey        — Sheet1 vertical header
C_TITLE_ORANGE = fill("FCE4D6")   # light orange — Sheet1 horizontal title
C_HDR_YELLOW   = fill("FFF2CC")   # yellow      — Sheet1 horizontal header
C_TITLE_TEAL   = fill("DDEBF7")   # teal-ish    — Sheet2 titles
C_HDR_BLUE2    = fill("DEEAF1")   # pale blue   — Sheet2 headers
C_HDR_GREEN    = fill("E2EFDA")   # green       — Sheet3 table1 header
C_HDR_PURPLE   = fill("EAD1DC")   # purple      — Sheet3 table2 header
C_HDR_TEAL2    = fill("D0E4F5")   # teal        — Sheet3 table3 header
C_HDR_PINK     = fill("FCE4D6")   # pink        — Sheet3 table4 header
C_HDR_LGREY    = fill("F2F2F2")   # light grey  — Sheet4 date headers
C_PRESENT      = fill("00B050")   # green       — present
C_ABSENT       = fill("FF0000")   # red         — absent

NAMES = [
    "Alice Chen", "Bob Smith", "Carol White", "David Lee", "Eva Brown",
    "Frank Kim", "Grace Liu", "Henry Park", "Iris Wang", "Jack Zhang",
    "Karen Ng", "Leo Tan", "Mia Patel", "Noah Singh", "Olivia Roy",
    "Peter Wu", "Quinn Adams", "Rachel Scott", "Sam Torres", "Tina Nguyen",
    "Uma Patel", "Victor Reyes", "Wendy Hall", "Xander Brooks", "Yuki Sato",
]
DEPTS      = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Operations"]
STATUSES   = ["Not Started", "In Progress", "Done", "Blocked", "Review"]
PRIORITIES = ["Critical", "High", "Medium", "Low"]
CATEGORIES = ["Salaries", "Software", "Hardware", "Travel", "Marketing", "Training"]
RISK_CATS  = ["Technical", "Financial", "Operational", "Legal", "Reputational"]
PRODUCTS   = ["Laptop Pro", "Wireless Mouse", "USB Hub", "Monitor 27\"", "Keyboard Mech",
              "Webcam HD", "Headset BT", "Desk Lamp", "Chair Ergo", "Standing Desk"]
SUBJECTS   = ["Login issue", "Payment failed", "Account locked", "Data export error",
              "Slow performance", "Feature request", "Billing query", "API timeout",
              "UI bug", "Integration broken"]

wb = openpyxl.Workbook()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Mixed Layout
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Mixed Layout"

# ── Vertical table: Employee Performance (A1:F52) ────────────────────────────
# Title: plain text, light-blue bg, NO bold, NO merge
ws1["A1"] = "Employee Performance Q1 2025"
ws1["A1"].fill = C_TITLE_BLUE

# Header: plain text, grey bg, NO bold
v_headers = ["Employee ID", "Name", "Department", "Score", "Bonus", "Review Date"]
for col, h in enumerate(v_headers, 1):
    ws1.cell(row=2, column=col, value=h).fill = C_HDR_GREY

for i in range(50):
    score = round(random.uniform(60, 100), 1)
    bonus = round(score * random.uniform(10, 25), 2)
    review = date(2025, 1, 1) + timedelta(days=random.randint(0, 89))
    ws1.append([f"EMP-{1000+i}", NAMES[i % len(NAMES)], DEPTS[i % len(DEPTS)],
                score, bonus, review])

# ── Horizontal table: Monthly KPIs (H1:R22) ──────────────────────────────────
# Title: plain text, orange bg, NO bold, NO merge
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]
H = 8  # col H

ws1.cell(row=1, column=H, value="Sales KPIs 2025").fill = C_TITLE_ORANGE

# Header row: "Metric" + month names, yellow bg, NO bold
ws1.cell(row=2, column=H, value="Metric").fill = C_HDR_YELLOW
for j, m in enumerate(MONTHS):
    ws1.cell(row=2, column=H+1+j, value=m).fill = C_HDR_YELLOW

KPI_METRICS = [
    "Revenue ($K)", "Units Sold", "New Customers", "Churn Rate (%)",
    "Avg Deal Size ($)", "Pipeline ($K)", "Win Rate (%)", "CAC ($)",
    "NPS Score", "Support Tickets", "Upsell Revenue ($K)",
    "Refund Rate (%)", "Active Users", "Conversion Rate (%)",
    "Gross Margin (%)", "ARPU ($)", "MRR ($K)", "Quota Attainment (%)",
    "Leads Generated", "Opportunities Created",
]
for row_off, metric in enumerate(KPI_METRICS):
    r = 3 + row_off
    ws1.cell(row=r, column=H, value=metric)
    for j in range(len(MONTHS)):
        ws1.cell(row=r, column=H+1+j, value=round(random.uniform(10, 500), 1))

# ── Stray / accidental text ───────────────────────────────────────────────────
for (r, c, txt) in [
    (55, 2, "TODO: verify this"), (57, 4, "check with manager"),
    (60, 1, "N/A data missing"),  (62, 5, "DRAFT"),
    (55, 9, "see email thread"),  (58, 11, "pending approval"),
    (63, 3, "old data ignore"),   (65, 7, "???"),
]:
    ws1.cell(row=r, column=c, value=txt)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Dense Tables (no row gaps between A and B, 1-col gap only)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Dense Tables")

# ── Table A: Project Tasks (A1:G31) ──────────────────────────────────────────
ws2["A1"] = "Project Tasks"
ws2["A1"].fill = C_TITLE_TEAL

a_hdrs = ["Task ID", "Task Name", "Assignee", "Status", "Priority", "Start Date", "End Date"]
for col, h in enumerate(a_hdrs, 1):
    ws2.cell(row=2, column=col, value=h).fill = C_HDR_BLUE2

for i in range(29):
    start = date(2025, 1, 1) + timedelta(days=random.randint(0, 60))
    end = start + timedelta(days=random.randint(1, 30))
    ws2.cell(row=3+i, column=1, value=f"TASK-{100+i}")
    ws2.cell(row=3+i, column=2, value=f"Task description number {i+1} for project alpha")
    ws2.cell(row=3+i, column=3, value=NAMES[i % len(NAMES)])
    ws2.cell(row=3+i, column=4, value=random.choice(STATUSES))
    ws2.cell(row=3+i, column=5, value=random.choice(PRIORITIES))
    ws2.cell(row=3+i, column=6, value=start)
    ws2.cell(row=3+i, column=7, value=end)

# ── Table B: Budget Tracker (I1:N31) — same row range, col H is the gap ──────
ws2["I1"] = "Budget Tracker"
ws2["I1"].fill = C_TITLE_TEAL

b_hdrs = ["Cost Center", "Category", "Budgeted", "Actual", "Variance", "Status"]
for col, h in enumerate(b_hdrs, 9):
    ws2.cell(row=2, column=col, value=h).fill = C_HDR_BLUE2

for i in range(29):
    budgeted = round(random.uniform(5000, 50000), 2)
    actual   = round(budgeted * random.uniform(0.7, 1.3), 2)
    ws2.cell(row=3+i, column=9,  value=f"CC-{200+i}")
    ws2.cell(row=3+i, column=10, value=random.choice(CATEGORIES))
    ws2.cell(row=3+i, column=11, value=budgeted)
    ws2.cell(row=3+i, column=12, value=actual)
    ws2.cell(row=3+i, column=13, value=round(actual - budgeted, 2))
    ws2.cell(row=3+i, column=14, value="Over" if actual > budgeted else "Under")

# ── Table C: Risk Register (A33:N63) — 1 empty row gap at row 32 ─────────────
ws2["A33"] = "Risk Register"
ws2["A33"].fill = C_TITLE_TEAL

c_hdrs = [
    "Risk ID", "Description", "Category", "Likelihood", "Impact",
    "Risk Score", "Owner", "Mitigation", "Status", "Review Date",
    "Residual Likelihood", "Residual Impact", "Residual Score", "Escalated",
]
for col, h in enumerate(c_hdrs, 1):
    ws2.cell(row=34, column=col, value=h).fill = C_HDR_BLUE2

for i in range(29):
    lik = random.randint(1, 5)
    imp = random.randint(1, 5)
    r_lik = max(1, lik - random.randint(0, 2))
    r_imp = max(1, imp - random.randint(0, 2))
    ws2.cell(row=35+i, column=1,  value=f"RISK-{300+i}")
    ws2.cell(row=35+i, column=2,  value=f"Risk scenario {i+1}: potential impact on deliverables")
    ws2.cell(row=35+i, column=3,  value=random.choice(RISK_CATS))
    ws2.cell(row=35+i, column=4,  value=lik)
    ws2.cell(row=35+i, column=5,  value=imp)
    ws2.cell(row=35+i, column=6,  value=lik * imp)
    ws2.cell(row=35+i, column=7,  value=NAMES[i % len(NAMES)])
    ws2.cell(row=35+i, column=8,  value=f"Mitigation plan {i+1}")
    ws2.cell(row=35+i, column=9,  value=random.choice(["Open", "Closed", "Monitoring"]))
    ws2.cell(row=35+i, column=10, value=date(2025, 2, 1) + timedelta(days=i*3))
    ws2.cell(row=35+i, column=11, value=r_lik)
    ws2.cell(row=35+i, column=12, value=r_imp)
    ws2.cell(row=35+i, column=13, value=r_lik * r_imp)
    ws2.cell(row=35+i, column=14, value="Yes" if lik * imp >= 16 else "No")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Random Scatter
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Random Scatter")

def place_table(ws, start_row, start_col, title, headers, data, title_fill, hdr_fill):
    """Write a titled table at arbitrary position, plain text, color-only signals."""
    ws.cell(row=start_row, column=start_col, value=title).fill = title_fill
    for j, h in enumerate(headers):
        ws.cell(row=start_row+1, column=start_col+j, value=h).fill = hdr_fill
    for ri, row in enumerate(data):
        for ci, val in enumerate(row):
            ws.cell(row=start_row+2+ri, column=start_col+ci, value=val)

# Table 1: Customer Orders (A3:H43)
order_data = []
for i in range(40):
    qty   = random.randint(1, 20)
    price = round(random.uniform(9.99, 999.99), 2)
    od    = date(2025, 1, 1) + timedelta(days=random.randint(0, 89))
    order_data.append([
        f"ORD-{5000+i}", NAMES[i % len(NAMES)], random.choice(PRODUCTS),
        qty, price, round(qty*price, 2),
        od, od + timedelta(days=random.randint(1, 7)),
    ])
place_table(ws3, 3, 1, "Customer Orders Q1 2025",
            ["Order ID", "Customer", "Product", "Qty", "Unit Price", "Total", "Order Date", "Ship Date"],
            order_data, C_TITLE_BLUE, C_HDR_GREEN)

# Table 2: Support Tickets (K3:P33)
ticket_data = []
for i in range(30):
    rd = date(2025, 1, 5) + timedelta(days=random.randint(0, 85)) if random.random() > 0.3 else None
    ticket_data.append([
        f"TKT-{7000+i}", random.choice(SUBJECTS), random.choice(PRIORITIES),
        NAMES[i % len(NAMES)], random.choice(["Open", "Closed", "Pending"]), rd,
    ])
place_table(ws3, 3, 11, "Support Tickets Q1 2025",
            ["Ticket ID", "Subject", "Priority", "Agent", "Status", "Resolved Date"],
            ticket_data, C_TITLE_ORANGE, C_HDR_PURPLE)

# Table 3: KPI Matrix horizontal (A47:I57) — metrics as rows, quarters as cols
QUARTERS = ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
            "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025"]
kpi_data = []
for metric in ["Revenue ($M)", "EBITDA ($M)", "Headcount", "Customer Count",
               "NPS", "Churn (%)", "CAC ($)", "LTV ($)", "Gross Margin (%)"]:
    kpi_data.append([metric] + [round(random.uniform(1, 1000), 2) for _ in QUARTERS])
place_table(ws3, 47, 1, "KPI Dashboard Quarterly",
            ["KPI Metric"] + QUARTERS,
            kpi_data, C_TITLE_TEAL, C_HDR_TEAL2)

# Table 4: Product Lookup (K47:N67)
lookup_data = []
for i in range(20):
    lookup_data.append([
        f"SKU-{9000+i}", random.choice(PRODUCTS),
        random.choice(["Electronics", "Furniture", "Accessories"]),
        round(random.uniform(9.99, 1999.99), 2),
    ])
place_table(ws3, 47, 11, "Product Catalog",
            ["SKU", "Product Name", "Category", "List Price"],
            lookup_data, C_TITLE_ORANGE, C_HDR_PINK)

# Stray text between tables
for (r, c, txt) in [
    (44, 3, "last updated Jan 2025"), (44, 12, "source CRM export"),
    (35, 9, "NOTE check duplicates"), (68, 2, "archived"), (68, 12, "v2.1"),
]:
    ws3.cell(row=r, column=c, value=txt)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Attendance (color-only grid)
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Attendance")

# 60 working days
work_days = []
d = date(2025, 1, 2)
while len(work_days) < 60:
    if d.weekday() < 5:
        work_days.append(d)
    d += timedelta(days=1)

# Header row: plain text, light grey bg, NO bold
ws4.cell(row=1, column=1, value="Name").fill = C_HDR_LGREY
for j, day in enumerate(work_days):
    c = ws4.cell(row=1, column=2+j, value=day.strftime("%d-%b"))
    c.fill = C_HDR_LGREY
    c.alignment = Alignment(horizontal="center", text_rotation=45)

# Data rows: name (plain, no bold) + color-only attendance cells
for i, member in enumerate(NAMES):
    ws4.cell(row=2+i, column=1, value=member)
    for j in range(len(work_days)):
        cell = ws4.cell(row=2+i, column=2+j, value="")
        cell.fill = C_PRESENT if random.random() < 0.85 else C_ABSENT

ws4.freeze_panes = "B2"
ws4.column_dimensions["A"].width = 18
for j in range(len(work_days)):
    ws4.column_dimensions[get_column_letter(2+j)].width = 5


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("stress_test.xlsx")
print("Created stress_test.xlsx")
print("  Sheet1 'Mixed Layout'  : vertical table (50r) + horizontal KPI (20 metrics x 10 months) + 8 stray cells — NO bold anywhere")
print("  Sheet2 'Dense Tables'  : Tasks + Budget (same row range, 1-col gap) + Risk Register (1-row gap) — color-only headers")
print("  Sheet3 'Random Scatter': 4 tables at scattered positions + 5 stray cells — color-only headers")
print("  Sheet4 'Attendance'    : 25 members x 60 days, green/red color-only data, plain text headers")
