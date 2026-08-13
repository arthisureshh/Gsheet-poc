"""
Generate complex_test_data.xlsx with 4 sheets for stress-testing
region detection, header inference, and chunking.

Sheet1 "Mixed Layout"   — vertical table + horizontal table side-by-side (no col gap),
                          plus accidental stray text in random cells
Sheet2 "Dense Tables"   — 3 tables with NO row gaps and NO column gaps (hardest case)
Sheet3 "Random Scatter" — tables at random positions, mixed orientations, large data
Sheet4 "Attendance"     — color-only grid (green=present, red=absent), no text in data cells

Run:
    python create_complex_test_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

random.seed(42)

GREEN = PatternFill("solid", fgColor="00B050")   # present
RED   = PatternFill("solid", fgColor="FF0000")   # absent
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
TITLE_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)

wb = openpyxl.Workbook()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Mixed Layout
# Vertical table (cols A-F, rows 1-52) + Horizontal table (cols H-R, rows 1-22)
# No column gap between them (col G is the only separator — but we'll make it
# a single empty col so region detector must find the gap).
# Stray text scattered in ~8 random cells outside both tables.
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Mixed Layout"

# ── Vertical table: Employee Performance (A1:F52) ────────────────────────────
ws1["A1"] = "Employee Performance — Q1 2025"
ws1.merge_cells("A1:F1")
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = TITLE_FILL
ws1["A1"].alignment = Alignment(horizontal="center")

v_headers = ["Employee ID", "Name", "Department", "Score", "Bonus ($)", "Review Date"]
for col, h in enumerate(v_headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

departments = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Operations"]
names = ["Alice Chen", "Bob Smith", "Carol White", "David Lee", "Eva Brown",
         "Frank Kim", "Grace Liu", "Henry Park", "Iris Wang", "Jack Zhang",
         "Karen Ng", "Leo Tan", "Mia Patel", "Noah Singh", "Olivia Roy"]

for i in range(50):
    emp = names[i % len(names)]
    dept = departments[i % len(departments)]
    score = round(random.uniform(60, 100), 1)
    bonus = round(score * random.uniform(10, 25), 2)
    review = date(2025, 1, 1) + timedelta(days=random.randint(0, 89))
    ws1.append([f"EMP-{1000+i}", emp, dept, score, bonus, review])

# ── Horizontal table: Monthly Sales KPIs (H1:R22) ────────────────────────────
# Headers in row 1 (col H = "Metric", cols I-R = Jan..Oct)
# Each subsequent row is a metric with 10 monthly values
months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]
h_start_col = 8   # col H

ws1.cell(row=1, column=h_start_col, value="Sales KPIs — 2025")
ws1.merge_cells(
    start_row=1, end_row=1,
    start_column=h_start_col, end_column=h_start_col + len(months)
)
title_cell = ws1.cell(row=1, column=h_start_col)
title_cell.font = TITLE_FONT
title_cell.fill = TITLE_FILL
title_cell.alignment = Alignment(horizontal="center")

ws1.cell(row=2, column=h_start_col, value="Metric").font = HEADER_FONT
for j, m in enumerate(months):
    c = ws1.cell(row=2, column=h_start_col + 1 + j, value=m)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL

metrics = [
    "Revenue ($K)", "Units Sold", "New Customers", "Churn Rate (%)",
    "Avg Deal Size ($)", "Pipeline ($K)", "Win Rate (%)", "CAC ($)",
    "NPS Score", "Support Tickets", "Upsell Revenue ($K)",
    "Refund Rate (%)", "Active Users", "Conversion Rate (%)",
    "Gross Margin (%)", "ARPU ($)", "MRR ($K)", "Quota Attainment (%)",
    "Leads Generated", "Opportunities Created",
]
for row_offset, metric in enumerate(metrics):
    r = 3 + row_offset
    ws1.cell(row=r, column=h_start_col, value=metric).font = HEADER_FONT
    for j in range(len(months)):
        ws1.cell(row=r, column=h_start_col + 1 + j,
                 value=round(random.uniform(10, 500), 1))

# ── Stray / accidental text ───────────────────────────────────────────────────
stray_positions = [(55, 2), (57, 4), (60, 1), (62, 5), (55, 9), (58, 11), (63, 3), (65, 7)]
stray_texts = [
    "TODO: verify this", "check with manager", "N/A — data missing",
    "DRAFT", "see email thread", "pending approval", "old data — ignore", "???"
]
for (r, c), txt in zip(stray_positions, stray_texts):
    ws1.cell(row=r, column=c, value=txt)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Dense Tables (no row gaps, no column gaps between tables)
# Three tables packed together:
#   Table A: Project Tasks     (rows 1-31,  cols A-G)
#   Table B: Budget Tracker    (rows 1-31,  cols I-N)   ← 1-col gap at H
#   Table C: Risk Register     (rows 33-63, cols A-N)   ← 1-row gap at 32
# The challenge: Table A and B share the same row range with only 1 col gap.
# Table C starts right after 1 empty row.
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Dense Tables")

# ── Table A: Project Tasks (A1:G31) ──────────────────────────────────────────
ws2["A1"] = "Project Tasks"
ws2.merge_cells("A1:G1")
ws2["A1"].font = TITLE_FONT; ws2["A1"].fill = TITLE_FILL

a_headers = ["Task ID", "Task Name", "Assignee", "Status", "Priority", "Start Date", "End Date"]
for col, h in enumerate(a_headers, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL

statuses = ["Not Started", "In Progress", "Done", "Blocked", "Review"]
priorities = ["Critical", "High", "Medium", "Low"]
for i in range(29):
    start = date(2025, 1, 1) + timedelta(days=random.randint(0, 60))
    end = start + timedelta(days=random.randint(1, 30))
    ws2.append([
        f"TASK-{100+i}",
        f"Task description number {i+1} for project alpha",
        names[i % len(names)],
        random.choice(statuses),
        random.choice(priorities),
        start, end,
    ])

# ── Table B: Budget Tracker (I1:N31) — same row range as Table A ─────────────
ws2["I1"] = "Budget Tracker"
ws2.merge_cells("I1:N1")
ws2["I1"].font = TITLE_FONT; ws2["I1"].fill = TITLE_FILL

b_headers = ["Cost Center", "Category", "Budgeted ($)", "Actual ($)", "Variance ($)", "Status"]
for col, h in enumerate(b_headers, 9):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL

categories = ["Salaries", "Software", "Hardware", "Travel", "Marketing", "Training"]
for i in range(29):
    budgeted = round(random.uniform(5000, 50000), 2)
    actual = round(budgeted * random.uniform(0.7, 1.3), 2)
    variance = round(actual - budgeted, 2)
    ws2.cell(row=3+i, column=9,  value=f"CC-{200+i}")
    ws2.cell(row=3+i, column=10, value=random.choice(categories))
    ws2.cell(row=3+i, column=11, value=budgeted)
    ws2.cell(row=3+i, column=12, value=actual)
    ws2.cell(row=3+i, column=13, value=variance)
    ws2.cell(row=3+i, column=14, value="Over" if variance > 0 else "Under")

# ── Table C: Risk Register (A33:N63) — starts after 1 empty row ──────────────
ws2["A33"] = "Risk Register"
ws2.merge_cells("A33:N33")
ws2["A33"].font = TITLE_FONT; ws2["A33"].fill = TITLE_FILL

c_headers = [
    "Risk ID", "Description", "Category", "Likelihood", "Impact",
    "Risk Score", "Owner", "Mitigation", "Status", "Review Date",
    "Residual Likelihood", "Residual Impact", "Residual Score", "Escalated"
]
for col, h in enumerate(c_headers, 1):
    c = ws2.cell(row=34, column=col, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL

risk_cats = ["Technical", "Financial", "Operational", "Legal", "Reputational"]
likelihoods = [1, 2, 3, 4, 5]
for i in range(29):
    lik = random.choice(likelihoods)
    imp = random.choice(likelihoods)
    r_lik = max(1, lik - random.randint(0, 2))
    r_imp = max(1, imp - random.randint(0, 2))
    ws2.cell(row=35+i, column=1,  value=f"RISK-{300+i}")
    ws2.cell(row=35+i, column=2,  value=f"Risk scenario {i+1}: potential impact on deliverables")
    ws2.cell(row=35+i, column=3,  value=random.choice(risk_cats))
    ws2.cell(row=35+i, column=4,  value=lik)
    ws2.cell(row=35+i, column=5,  value=imp)
    ws2.cell(row=35+i, column=6,  value=lik * imp)
    ws2.cell(row=35+i, column=7,  value=names[i % len(names)])
    ws2.cell(row=35+i, column=8,  value=f"Mitigation plan {i+1}")
    ws2.cell(row=35+i, column=9,  value=random.choice(["Open", "Closed", "Monitoring"]))
    ws2.cell(row=35+i, column=10, value=date(2025, 2, 1) + timedelta(days=i*3))
    ws2.cell(row=35+i, column=11, value=r_lik)
    ws2.cell(row=35+i, column=12, value=r_imp)
    ws2.cell(row=35+i, column=13, value=r_lik * r_imp)
    ws2.cell(row=35+i, column=14, value="Yes" if lik * imp >= 16 else "No")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Random Scatter
# 4 tables at non-adjacent positions with varying orientations and sizes.
# Table 1: top-left,  large vertical (A2:H42)
# Table 2: top-right, medium vertical (K2:P32)
# Table 3: bottom-left, horizontal KPI matrix (A46:M56)
# Table 4: bottom-right, small lookup table (K46:N66)
# Accidental text in 6 cells between tables.
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Random Scatter")

def write_table(ws, start_row, start_col, title, headers, data_rows):
    """Helper to write a titled table at an arbitrary position."""
    end_col = start_col + len(headers) - 1
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    if end_col > start_col:
        ws.merge_cells(
            start_row=start_row, end_row=start_row,
            start_column=start_col, end_column=end_col
        )
    for col_offset, h in enumerate(headers):
        c = ws.cell(row=start_row+1, column=start_col+col_offset, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for row_offset, row in enumerate(data_rows):
        for col_offset, val in enumerate(row):
            ws.cell(row=start_row+2+row_offset, column=start_col+col_offset, value=val)

# Table 1: Customer Orders (A2:H42)
order_headers = ["Order ID", "Customer", "Product", "Qty", "Unit Price ($)", "Total ($)", "Order Date", "Ship Date"]
order_data = []
products = ["Laptop Pro", "Wireless Mouse", "USB Hub", "Monitor 27\"", "Keyboard Mech",
            "Webcam HD", "Headset BT", "Desk Lamp", "Chair Ergo", "Standing Desk"]
for i in range(40):
    qty = random.randint(1, 20)
    price = round(random.uniform(9.99, 999.99), 2)
    odate = date(2025, 1, 1) + timedelta(days=random.randint(0, 89))
    sdate = odate + timedelta(days=random.randint(1, 7))
    order_data.append([
        f"ORD-{5000+i}", names[i % len(names)], random.choice(products),
        qty, price, round(qty * price, 2), odate, sdate
    ])
write_table(ws3, 2, 1, "Customer Orders — Q1 2025", order_headers, order_data)

# Table 2: Support Tickets (K2:P32)
ticket_headers = ["Ticket ID", "Subject", "Priority", "Agent", "Status", "Resolved Date"]
ticket_data = []
subjects = ["Login issue", "Payment failed", "Account locked", "Data export error",
            "Slow performance", "Feature request", "Billing query", "API timeout",
            "UI bug", "Integration broken"]
for i in range(30):
    rdate = date(2025, 1, 5) + timedelta(days=random.randint(0, 85)) if random.random() > 0.3 else None
    ticket_data.append([
        f"TKT-{7000+i}", random.choice(subjects), random.choice(priorities),
        names[i % len(names)], random.choice(["Open", "Closed", "Pending"]), rdate
    ])
write_table(ws3, 2, 11, "Support Tickets — Q1 2025", ticket_headers, ticket_data)

# Table 3: Horizontal KPI Matrix (A46:M56) — metrics as rows, quarters as cols
ws3.cell(row=46, column=1, value="KPI Dashboard — Quarterly").font = TITLE_FONT
ws3.cell(row=46, column=1).fill = TITLE_FILL
ws3.merge_cells(start_row=46, end_row=46, start_column=1, end_column=9)
ws3.cell(row=47, column=1, value="KPI Metric").font = HEADER_FONT
quarters = ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
            "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
            "YTD 2025", "Target 2025", "Variance", "% Achieved"]
for j, q in enumerate(quarters):
    c = ws3.cell(row=47, column=2+j, value=q)
    c.font = HEADER_FONT; c.fill = HEADER_FILL

kpi_metrics = ["Revenue ($M)", "EBITDA ($M)", "Headcount", "Customer Count",
               "NPS", "Churn (%)", "CAC ($)", "LTV ($)", "Gross Margin (%)"]
for i, metric in enumerate(kpi_metrics):
    ws3.cell(row=48+i, column=1, value=metric).font = HEADER_FONT
    for j in range(12):
        ws3.cell(row=48+i, column=2+j, value=round(random.uniform(1, 1000), 2))

# Table 4: Product Lookup (K46:N66)
lookup_headers = ["SKU", "Product Name", "Category", "List Price ($)"]
lookup_data = []
for i in range(20):
    lookup_data.append([
        f"SKU-{9000+i}", random.choice(products),
        random.choice(["Electronics", "Furniture", "Accessories"]),
        round(random.uniform(9.99, 1999.99), 2)
    ])
write_table(ws3, 46, 11, "Product Catalog", lookup_headers, lookup_data)

# Accidental stray text between tables
stray3 = [(44, 3, "last updated: Jan 2025"), (44, 12, "source: CRM export"),
          (35, 9, "NOTE: check duplicates"), (68, 2, "archived"), (68, 12, "v2.1")]
for r, c, txt in stray3:
    ws3.cell(row=r, column=c, value=txt)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Attendance (color-only grid)
# Row headers = team member names (col A)
# Column headers = dates (row 1)
# Data cells: GREEN fill = present, RED fill = absent
# No text in data cells — purely color-coded
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Attendance")

team = [
    "Alice Chen", "Bob Smith", "Carol White", "David Lee", "Eva Brown",
    "Frank Kim", "Grace Liu", "Henry Park", "Iris Wang", "Jack Zhang",
    "Karen Ng", "Leo Tan", "Mia Patel", "Noah Singh", "Olivia Roy",
    "Peter Wu", "Quinn Adams", "Rachel Scott", "Sam Torres", "Tina Nguyen",
    "Uma Patel", "Victor Reyes", "Wendy Hall", "Xander Brooks", "Yuki Sato",
]

# 60 working days starting 2025-01-02
work_days = []
d = date(2025, 1, 2)
while len(work_days) < 60:
    if d.weekday() < 5:   # Mon-Fri
        work_days.append(d)
    d += timedelta(days=1)

# Header row: "Name" + dates
ws4.cell(row=1, column=1, value="Name").font = HEADER_FONT
ws4.cell(row=1, column=1).fill = HEADER_FILL
for j, day in enumerate(work_days):
    c = ws4.cell(row=1, column=2+j, value=day)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.number_format = "DD-MMM"
    c.alignment = Alignment(horizontal="center", text_rotation=45)

# Data rows: name + color cells
for i, member in enumerate(team):
    ws4.cell(row=2+i, column=1, value=member).font = Font(bold=True)
    for j in range(len(work_days)):
        # ~85% attendance rate, with some members having streaks
        present = random.random() < 0.85
        cell = ws4.cell(row=2+i, column=2+j, value="")
        cell.fill = GREEN if present else RED

# Freeze panes so name column and date row stay visible
ws4.freeze_panes = "B2"

# Column widths
ws4.column_dimensions["A"].width = 18
for j in range(len(work_days)):
    ws4.column_dimensions[get_column_letter(2+j)].width = 5


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("complex_test_data.xlsx")
print("Created complex_test_data.xlsx")
print("  Sheet1 'Mixed Layout'  : vertical Employee table (50 rows) + horizontal Sales KPI table (20 metrics x 10 months) + 8 stray cells")
print("  Sheet2 'Dense Tables'  : 3 tables — Project Tasks + Budget Tracker (same row range, 1-col gap) + Risk Register (1-row gap)")
print("  Sheet3 'Random Scatter': 4 tables at scattered positions — Orders(40r), Tickets(30r), KPI Matrix(9x12), Product Lookup(20r)")
print("  Sheet4 'Attendance'    : 25 team members x 60 work days, color-only (green=present, red=absent)")
