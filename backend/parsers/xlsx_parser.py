import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple
from backend.models import ParsedSheet, CellGrid, CellMeta, MetaGrid

_NONE_COLORS = {"00000000", "FFFFFFFF", "FF000000", "FFFFFFFF", None}


def _norm_color(raw) -> str | None:
    """Extract hex RRGGBB from openpyxl color, return None for transparent/white/black defaults."""
    if raw is None:
        return None
    try:
        if raw.type == "rgb":
            argb = raw.rgb  # 8-char AARRGGBB
            if argb in _NONE_COLORS:
                return None
            return argb[2:]  # strip alpha → RRGGBB
        if raw.type == "theme":
            return f"theme:{raw.theme}"
    except Exception:
        pass
    return None


_MEANINGFUL_COLORS = {"00B050", "FF0000"}  # green=present, red=absent — treat as data


def _cell_effective_value(cell) -> object:
    """Return cell value, or indexedText-style sentinel for ANY meaningful fill color.
    Mirrors Paxi buildIndexedCellText: color-only cell -> '[bg=#RRGGBB]'
    """
    if cell.value is not None and str(cell.value).strip():
        return cell.value
    if cell.fill and cell.fill.patternType not in (None, "none"):
        color = _norm_color(cell.fill.fgColor)
        if color and not color.startswith("theme:"):  # any explicit RGB color
            return f"[bg=#{color.upper()}]"
    return cell.value


def parse_xlsx(file_path: str) -> list[ParsedSheet]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = []
    try:
        for ws in wb.worksheets:
            rows = [[_cell_effective_value(cell) for cell in row] for row in ws.iter_rows()]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            grid = CellGrid(grid=rows, row_count=max_row, col_count=max_col)

            meta_rows: list[list[CellMeta]] = []
            for row in ws.iter_rows():
                meta_row: list[CellMeta] = []
                for cell in row:
                    bold = bool(cell.font and cell.font.bold)
                    bg = _norm_color(cell.fill.fgColor) if cell.fill and cell.fill.patternType not in (None, "none") else None
                    fc = _norm_color(cell.font.color) if cell.font and cell.font.color else None
                    meta_row.append(CellMeta(bold=bold, bg_color=bg, font_color=fc))
                meta_rows.append(meta_row)
            meta_grid = MetaGrid(grid=meta_rows, row_count=max_row, col_count=max_col)

            merged = [
                {"min_row": r.min_row, "max_row": r.max_row, "min_col": r.min_col, "max_col": r.max_col}
                for r in ws.merged_cells.ranges
            ]
            fp = ws.freeze_panes
            if fp:
                freeze_row = coordinate_to_tuple(fp)[0] - 1 if isinstance(fp, str) else fp.row - 1
            else:
                freeze_row = 0

            sheets.append(ParsedSheet(
                sheet_name=ws.title,
                rows=rows,
                grid=grid,
                meta_grid=meta_grid,
                merged_ranges=merged,
                freeze_row=freeze_row,
                max_col=max_col,
            ))
    finally:
        wb.close()
    return sheets
