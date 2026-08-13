"""
Generate complex_test_data-v2.xlsx by copying V1 and applying targeted changes.
NEVER deletes rows — blanks them instead to preserve row positions and region detection.
"""
import shutil
import openpyxl
from datetime import date

shutil.copy('complex_test_data.xlsx', 'complex_test_data-v2.xlsx')
wb = openpyxl.load_workbook('complex_test_data-v2.xlsx')


# ── Sheet1: Mixed Layout — Employee Performance (rows 3..52, cols A-F) ────────
ws1 = wb['Mixed Layout']

rows_to_blank = []
for row in ws1.iter_rows(min_row=3, max_col=6):
    val = row[0].value
    if val in ('EMP-1004', 'EMP-1005', 'EMP-1006'):
        rows_to_blank.append(row[0].row)
    if val == 'EMP-1009':
        ws1.cell(row[0].row, 4).value = 99.9
        ws1.cell(row[0].row, 5).value = 1500.0

# Replace the 3 deleted employees with 3 new ones in-place
new_emps = [
    ('EMP-1050', 'Zara Ahmed', 'Engineering', 87.5, 1450.0, date(2025, 3,  1)),
    ('EMP-1051', 'Liam Chen',  'Sales',       76.2,  980.0, date(2025, 3,  6)),
    ('EMP-1052', 'Nora Kim',   'Finance',     91.0, 1820.0, date(2025, 3, 11)),
]
for i, r in enumerate(rows_to_blank):
    eid, name, dept, score, bonus, rev = new_emps[i]
    ws1.cell(r, 1).value = eid
    ws1.cell(r, 2).value = name
    ws1.cell(r, 3).value = dept
    ws1.cell(r, 4).value = score
    ws1.cell(r, 5).value = bonus
    ws1.cell(r, 6).value = rev


# ── Sheet2: Dense Tables — Project Tasks (rows 3..31, cols A-G) ───────────────
ws2 = wb['Dense Tables']

task_row_to_blank = None
for row in ws2.iter_rows(min_row=3, max_row=31, max_col=7):
    val = row[0].value
    if val == 'TASK-110':
        task_row_to_blank = row[0].row
    if val == 'TASK-103':
        ws2.cell(row[0].row, 4).value = 'Done'

# Replace TASK-110 with TASK-129, use empty gap row 32 for TASK-130
new_tasks = [
    ('TASK-129', 'New feature: dashboard export', 'Alice Chen', 'Not Started', 'High',     date(2025, 3, 10), date(2025, 3, 25)),
    ('TASK-130', 'Bug fix: login timeout issue',  'Bob Smith',  'In Progress', 'Critical', date(2025, 3, 12), date(2025, 3, 15)),
]
target_rows = [task_row_to_blank, 32]  # row 32 is the empty gap between Tasks and Risk Register
for (tid, tname, assignee, status, priority, start, end), r in zip(new_tasks, target_rows):
    for col, val in enumerate([tid, tname, assignee, status, priority, start, end], 1):
        ws2.cell(r, col).value = val


# ── Sheet3: Random Scatter — Customer Orders (rows 4..43, cols A-H) ───────────
ws3 = wb['Random Scatter']

ord5015_row = None
for row in ws3.iter_rows(min_row=4, max_row=43, max_col=8):
    val = row[0].value
    if val == 'ORD-5015':
        ord5015_row = row[0].row
    if val == 'ORD-5002':
        ws3.cell(row[0].row, 4).value = 99
        ws3.cell(row[0].row, 6).value = round(99 * ws3.cell(row[0].row, 5).value, 2)

# Replace ORD-5015 with ORD-5040, use row 44 (stray text row) for ORD-5041
new_orders = [
    ('ORD-5040', 'Zara Ahmed', 'Standing Desk', 2, 499.99, 999.98, date(2025, 3, 20), date(2025, 3, 22)),
    ('ORD-5041', 'Liam Chen',  'Webcam HD',     5,  89.99, 449.95, date(2025, 3, 21), date(2025, 3, 23)),
]
# Find the stray text row after orders (row 44 in V1)
stray_row = next(r for r in range(43, 47) if ws3.cell(r, 3).value == 'last updated: Jan 2025')
target_rows = [ord5015_row, stray_row]
for (oid, customer, product, qty, price, total, odate, sdate), r in zip(new_orders, target_rows):
    for c in range(1, 9):
        ws3.cell(r, c).value = None
    ws3.cell(r, 1).value = oid
    ws3.cell(r, 2).value = customer
    ws3.cell(r, 3).value = product
    ws3.cell(r, 4).value = qty
    ws3.cell(r, 5).value = price
    ws3.cell(r, 6).value = total
    ws3.cell(r, 7).value = odate
    ws3.cell(r, 8).value = sdate


wb.save('complex_test_data-v2.xlsx')
print('Created complex_test_data-v2.xlsx')
print('  Sheet1: EMP-1004/1005/1006 replaced with EMP-1050/1051/1052, EMP-1009 edited')
print('  Sheet2: TASK-110 replaced with TASK-129, TASK-130 in gap row, TASK-103 status->Done')
print('  Sheet3: ORD-5015 replaced with ORD-5040, ORD-5041 in stray row, ORD-5002 qty->99')
print('  Sheet4: unchanged')
