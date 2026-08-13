import openpyxl
wb = openpyxl.load_workbook('complex_test_data-v2.xlsx')

ws1 = wb['Mixed Layout']
print('=== Sheet1 ===')
print('max_row:', ws1.max_row)
# Find last EMP row
emp_rows = [r for r in range(3, ws1.max_row+1) if ws1.cell(r,1).value and str(ws1.cell(r,1).value).startswith('EMP-')]
print('EMP rows:', emp_rows[-5:])
# Check what's around col H (KPI table)
print('KPI col H rows 1-5:', [(r, ws1.cell(r,8).value) for r in range(1,6)])
print('merged:', [str(m) for m in ws1.merged_cells.ranges][:5])

ws3 = wb['Random Scatter']
print('\n=== Sheet3 ===')
print('max_row:', ws3.max_row)
ord_rows = [r for r in range(1, ws3.max_row+1) if ws3.cell(r,1).value and str(ws3.cell(r,1).value).startswith('ORD-')]
print('ORD rows:', ord_rows[-5:])
print('Row 44 col1:', ws3.cell(44,1).value)
print('Row 45 col1:', ws3.cell(45,1).value)
print('Row 46 col1:', ws3.cell(46,1).value)
