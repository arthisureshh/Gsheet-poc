"""
Run once to generate test_data.xlsx with two tables on one sheet
and a second sheet with a CSV-style table.

Usage:
    python create_test_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date

wb = openpyxl.Workbook()

# ── Sheet 1: two tables separated by an empty row ──────────────────────────
ws1 = wb.active
ws1.title = "Project Tracker"

# Table 1 title (full-width merge = region boundary signal)
ws1.append(["Project Tracker — Sprint 3"])
ws1.merge_cells("A1:F1")
ws1["A1"].font = Font(bold=True)

# Table 1 headers + data
ws1.append(["Item", "Owner", "Status", "Due Date", "Points", "Sprint"])
rows1 = [
    ["Build login page",   "Alice", "Done",    date(2025, 1, 10), 5,  "Sprint 3"],
    ["Fix auth bug",       "Bob",   "Active",  date(2025, 1, 15), 3,  "Sprint 3"],
    ["Write unit tests",   "Alice", "Overdue", date(2025, 1, 8),  4,  "Sprint 3"],
    ["Deploy to staging",  "Carol", "Active",  date(2025, 1, 20), 8,  "Sprint 3"],
    ["Update docs",        "Bob",   "Overdue", date(2025, 1, 5),  2,  "Sprint 3"],
    ["Code review PR #12", "Carol", "Done",    date(2025, 1, 12), 3,  "Sprint 3"],
]
for r in rows1:
    ws1.append(r)

# Empty row = boundary between tables
ws1.append([])

# Table 2 title
ws1.append(["Bug Report Log"])
ws1.merge_cells("A10:F10")
ws1["A10"].font = Font(bold=True)

# Table 2 headers + data
ws1.append(["Bug ID", "Description", "Severity", "Reporter", "Status", "Resolved"])
rows2 = [
    ["BUG-001", "Login timeout",       "High",   "Alice", "Open",     None],
    ["BUG-002", "CSV export broken",   "Medium", "Bob",   "Resolved", date(2025, 1, 9)],
    ["BUG-003", "UI misalignment",     "Low",    "Carol", "Open",     None],
    ["BUG-004", "Null pointer in API", "High",   "Alice", "Open",     None],
    ["BUG-005", "Slow query on load",  "Medium", "Bob",   "Resolved", date(2025, 1, 11)],
]
for r in rows2:
    ws1.append(r)

# ── Sheet 2: inventory table ────────────────────────────────────────────────
ws2 = wb.create_sheet("Inventory")
ws2.append(["Product", "Category", "Stock", "Price", "Last Updated"])
inventory = [
    ["Widget A",  "Electronics", 120, 29.99,  date(2025, 1, 1)],
    ["Widget B",  "Electronics", 45,  49.99,  date(2025, 1, 3)],
    ["Gadget X",  "Accessories", 200, 9.99,   date(2025, 1, 2)],
    ["Gadget Y",  "Accessories", 0,   14.99,  date(2025, 1, 5)],
    ["Tool Z",    "Hardware",    88,  99.99,  date(2025, 1, 4)],
    ["Part Q",    "Hardware",    15,  5.49,   date(2025, 1, 6)],
]
for r in inventory:
    ws2.append(r)

wb.save("test_data.xlsx")
print("Created test_data.xlsx")
print("  Sheet 1 'Project Tracker': 2 tables (sprint items + bug log)")
print("  Sheet 2 'Inventory': 1 table (product inventory)")
