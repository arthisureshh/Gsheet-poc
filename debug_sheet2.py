import openpyxl
wb = openpyxl.load_workbook('complex_test_data.xlsx')
ws2 = wb['Dense Tables']
task110_row = next(r for r in range(3,32) if ws2.cell(r,1).value == 'TASK-110')
print('TASK-110 at row:', task110_row)
ws2.delete_rows(task110_row)
last_task = max(r for r in range(3,35) if ws2.cell(r,1).value and str(ws2.cell(r,1).value).startswith('TASK-'))
print('last_task after delete:', last_task)
print('merged after delete:', [str(m) for m in ws2.merged_cells.ranges])
r1, r2 = last_task+1, last_task+2
print(f'target rows: {r1}, {r2}')
for r in [r1, r2]:
    for c in range(1, 8):
        cell = ws2.cell(r, c)
        print(f'  row{r} col{c}: type={type(cell).__name__} val={cell.value}')
